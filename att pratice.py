import tkinter as tk
from tkinter import messagebox
import openpyxl
from datetime import datetime
import os

class EmployeeApp:
    def __init__(self, master):
        self.master = master
        self.entry_var = tk.StringVar()
        self.start_time = None
        self.wb = None
        self.current_date = None
        self.current_sheet = None

        self.create_workbook_if_not_exists()
        self.load_workbook()

        self.create_widgets()

    def create_workbook_if_not_exists(self):
        filename = 'new_employee_data.xlsx'
        if not os.path.isfile(filename):
            wb = openpyxl.Workbook()
            wb.save(filename)

    def load_workbook(self):
        filename = 'new_employee_data.xlsx'
        try:
            self.wb = openpyxl.load_workbook(filename)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load workbook: {e}")

    def create_widgets(self):
        frame = tk.Frame(self.master)
        frame.pack(padx=10, pady=10)

        self.entry = tk.Entry(frame, textvariable=self.entry_var)
        self.entry.grid(row=0, column=0, padx=5, pady=5)

        tk.Button(frame, text="Enter", command=self.enter_button_clicked).grid(row=1, column=0, padx=5, pady=5)

        self.emp_id_input = tk.Label(frame, text="")
        self.emp_id_input.grid(row=1, column=1, padx=5, pady=5)

        tk.Button(frame, text="Start", command=self.record_start_time).grid(row=2, column=0, padx=5, pady=5)
        tk.Button(frame, text="End", command=self.record_end_time).grid(row=2, column=1, padx=5, pady=5)

    def enter_button_clicked(self):
        employee_id = self.entry_var.get().strip()
        if employee_id:
            self.emp_id_input.config(text="Employee ID: " + employee_id)
            self.entry.config(state=tk.DISABLED)
        else:
            messagebox.showwarning("Invalid Input", "Please enter a valid Employee ID.")

    def record_start_time(self):
        try:
            employee_id = self.entry_var.get()

            if not employee_id:
                messagebox.showwarning("Missing Employee ID", "Please enter Employee ID.")
                return

            if self.wb is None:
                self.load_workbook()

            if self.current_date is None:
                self.current_date = datetime.now().date()
                self.create_new_sheet_if_not_exists()

            if self.current_sheet is None:
                messagebox.showerror("Sheet Error", "No active sheet found.")
                return

            row = None
            for idx, each_row in enumerate(self.current_sheet.iter_rows(values_only=True)):
                if each_row[0] == employee_id:
                    row = idx + 1
                    break

            if row is None:
                row = self.current_sheet.max_row + 1
                self.current_sheet.cell(row=row, column=1).value = employee_id

            self.start_time = datetime.now()
            self.current_sheet.cell(row=row, column=2).value = self.start_time.strftime("%H:%M:%S")

            self.wb.save("new_employee_data.xlsx")
            messagebox.showinfo("Start Time Recorded", f"Start time recorded for Employee ID: {employee_id}")
        except Exception as e:
            messagebox.showerror("Error", f"Error occurred while recording start time: {e}")

    def record_end_time(self):
        try:
            employee_id = self.entry_var.get()

            if not employee_id:
                messagebox.showwarning("Missing Employee ID", "Please enter Employee ID.")
                return

            if self.wb is None:
                self.load_workbook()

            if self.current_date is None:
                self.current_date = datetime.now().date()
                self.create_new_sheet_if_not_exists()

            if self.current_sheet is None:
                messagebox.showerror("Sheet Error", "No active sheet found.")
                return

            row = None
            for idx, each_row in enumerate(self.current_sheet.iter_rows(values_only=True)):
                if each_row[0] == employee_id:
                    row = idx + 1
                    break

            if row is None:
                messagebox.showinfo("Error", f"There's no start time record for Employee ID: {employee_id}.")
                return

            end_time = datetime.now()
            elapsed_time = end_time - self.start_time

            self.current_sheet.cell(row=row, column=3).value = end_time.strftime("%H:%M:%S")
            self.current_sheet.cell(row=row, column=4).value = str(elapsed_time)

            self.wb.save("new_employee_data.xlsx")
            messagebox.showinfo("End Time Recorded", f"End time recorded for Employee ID: {employee_id}")
        except Exception as e:
            messagebox.showerror("Error", f"Error occurred while recording end time: {e}")

    def create_new_sheet_if_not_exists(self):
        if self.current_sheet is None:
            try:
                self.current_sheet = self.wb.create_sheet(title=self.current_date.strftime("%Y-%m-%d"))
                self.current_sheet.append(["Employee ID", "Start Time", "End Time", "Elapsed Time"])
            except Exception as e:
                messagebox.showerror("Error", f"Error occurred while creating new sheet: {e}")

def main():
    root = tk.Tk()
    app = EmployeeApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()


