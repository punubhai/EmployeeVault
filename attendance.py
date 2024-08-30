import tkinter as tk
from tkinter import messagebox
import sqlite3
from datetime import datetime
from openpyxl import Workbook,load_workbook
import main

class EmployeeDatabaseApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Employee Database")
        self.master.geometry('1500x999')


        # Database connection
        self.conn = sqlite3.connect("employees.db")

        # Create GUI elements
        self.emp_id_label = tk.Label(master, text="Employee ID:")
        self.emp_id_label.grid(row=0, column=0, padx=10, pady=5)

        self.background_image = tk.PhotoImage(file="7.png")  # Change the file name to your image file
        self.background_label = tk.Label(master, image=self.background_image)
        self.background_label.place(x=0, y=0, relwidth=1, relheight=1)

        self.emp_id_entry = tk.Entry(master)
        self.emp_id_entry.place(x=660,y=210,height=30,width=140)

        self.check_button = tk.Button(master, text="Check", command=self.check_employee,background='#C8052F')
        self.check_button.place(x=810,y=215,width=160)

        self.result_label = tk.Label(master, text="", background='#C8052F', foreground='white', font=('bold', 25))
        self.result_label.place(x=700, y=290)

        # Adding text field and button on the left side


        self.text_field_entry = tk.Entry(master, state=tk.DISABLED,background='#810007')
        self.text_field_entry.place(x=280,y=670,width=160,height=20)

        self.enter_button = tk.Button(master, text="Enter", command=self.enter_text, state=tk.DISABLED,background='#810007')
        self.enter_button.place(x=280,y=710,width=160)

        # Adding text field and button on the right side

        self.exit_text_field_entry = tk.Entry(master, state=tk.DISABLED,background='#810007')
        self.exit_text_field_entry.place(x=1180,y=680,width=160,height=20)

        self.exit_button = tk.Button(master, text="Exit", command=self.exit_employee, state=tk.DISABLED,background='#810007')
        self.exit_button.place(x=1180,y=710,width=160,height=20)

        self.back = tk.Button(master, text="back", command=self.back, background='#463E32',foreground='white')
        self.back.place(x=10,y=100,width=160)

        # Store entry and exit time
        self.entry_time = None
        self.exit_time = None

    def check_employee(self):
        emp_id = self.emp_id_entry.get()

        if not emp_id:
            messagebox.showerror("Error", "Please enter an employee ID.")
            return

        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM employees WHERE employee_id=?", (emp_id,))
        employee = cursor.fetchone()

        if employee:
            # Display employee details
            self.result_label.config(text=f"WELCOME: {employee[1]}\n")
            # Enable the text field and the Enter button
            self.text_field_entry.config(state=tk.NORMAL)
            self.enter_button.config(state=tk.NORMAL)
            self.exit_text_field_entry.config(state=tk.NORMAL)
            self.exit_button.config(state=tk.NORMAL)
            # Record entry time
            self.entry_time = datetime.now()

        else:
            self.result_label.config(text="Employee not found.")
            # Disable the text field and the Enter button
            self.text_field_entry.config(state=tk.DISABLED)
            self.enter_button.config(state=tk.DISABLED)

    def enter_text(self):
        entered_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.text_field_entry.config(state=tk.NORMAL)
        self.text_field_entry.delete(0, tk.END)
        self.text_field_entry.insert(0, entered_text)
        self.text_field_entry.config(state=tk.DISABLED)

    def exit_employee(self):
        # Record exit time
        entered_text = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.exit_text_field_entry.config(state=tk.NORMAL)
        self.exit_text_field_entry.delete(0, tk.END)
        self.exit_text_field_entry.insert(0, entered_text)
        self.exit_text_field_entry.config(state=tk.DISABLED)
        # Calculate time difference
        self.exit_time = datetime.now()
        time_difference = self.exit_time - self.entry_time
        # Display time difference
        messagebox.showinfo("Time Difference", f"Time difference: {time_difference}")
        # Save data to Excel
        self.save_to_excel()



    def save_to_excel(self):
        emp_id = self.emp_id_entry.get()
        entry_time = self.entry_time.strftime("%H:%M:%S")  # Extract only time
        exit_time = self.exit_time.strftime("%H:%M:%S")    # Extract only time
        time_difference = str(self.exit_time - self.entry_time)

        # Create or load the main Excel file
        filename = "employee_data_all.xlsx"
        try:
            wb = load_workbook(filename)
        except FileNotFoundError:
            # If the file doesn't exist, create a new one
            wb = Workbook()
            ws = wb.active
            ws.title = "Permanent Data"
            ws.append(['Employee ID', 'Entry Time', 'Exit Time', 'Time Spent'])
        else:
            ws = wb.active

        # Append data to the main sheet
        ws.append([emp_id, entry_time, exit_time, time_difference])

        # Create or load the date-specific sheet
        date_sheet_name = datetime.now().strftime('%Y-%m-%d')
        if date_sheet_name not in wb.sheetnames:
            ws_date = wb.create_sheet(title=date_sheet_name)
            ws_date.append(['Employee ID', 'Entry Time', 'Exit Time', 'Time Spent'])
        else:
            ws_date = wb[date_sheet_name]

        # Append data to the date-specific sheet
        ws_date.append([emp_id, entry_time, exit_time, time_difference])

        # Save the Excel file
        wb.save(filename)

        # Show info message
        messagebox.showinfo("Excel File Saved", f"Employee data saved to {filename}")


    def back(self):
         self.master.destroy()  # Destroy current window
         new_root = tk.Tk()  # Create a new Tkinter window
         main.App(new_root)  # Initialize the next page
         new_root.mainloop()



